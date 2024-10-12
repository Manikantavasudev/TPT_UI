import tkinter as tk
from tkinter import filedialog, messagebox
import json
import openpyxl
from openpyxl.styles import PatternFill

# Define colors for results
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for Pass
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red for Fail
pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")   # Pink for Inconclusive
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for Missing Tests

def upload_json_file(label):
    filepath = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if filepath:
        label.config(text=filepath)
    return filepath

def compare_jsons(golden_json_path, report_json_path):
    try:
        with open(golden_json_path, 'r') as golden_file, open(report_json_path, 'r') as report_file:
            golden_data = json.load(golden_file)
            report_data = json.load(report_file)
    except FileNotFoundError:
        messagebox.showerror("Error", "File not found. Please upload valid files.")
        return
    except json.JSONDecodeError:
        messagebox.showerror("Error", "Invalid JSON format. Please check your files.")
        return

    test_info = []
    
    golden_tests = golden_data.get('TestingScope', [])
    report_tests = report_data.get('TestingScope', [])

    report_test_dict = {test.get('TestId'): test.get('TestResult') for test in report_tests}
    
    if len(golden_tests) != len(report_tests):
        messagebox.showwarning("Warning", "Test count mismatch between Golden report and uploaded report.")

    for idx, golden_test in enumerate(golden_tests, start=1):
        test_id_golden = golden_test.get('TestId')
        test_result_golden = golden_test.get('TestResult')
        
        test_result_report = report_test_dict.get(test_id_golden, None)
        
        comment = ""
        if test_result_report is None:
            comment = "Test didn't execute"
            test_result_report = "Missing"
        elif test_result_golden != test_result_report:
            comment = f"Test Result mismatch: Golden: {test_result_golden}, Report: {test_result_report}"
        
        test_info.append((idx, test_id_golden, test_result_golden, test_result_report, comment))
    
    write_comparison_to_excel(test_info)

def write_comparison_to_excel(test_info, filename="ComparisonResults.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set headers
    ws.append(["SN", "TestId", "Golden TestResult", "Report TestResult", "Comments"])

    # Add data rows and apply colors based on results
    for row in test_info:
        ws.append(row)
        row_num = ws.max_row
        
        # Apply color based on Report TestResult
        report_result_cell = ws.cell(row=row_num, column=4)
        report_result = report_result_cell.value.lower()

        if report_result == "pass":
            report_result_cell.fill = green_fill
        elif report_result == "fail":
            report_result_cell.fill = red_fill
        elif report_result == "inconclusive":
            report_result_cell.fill = pink_fill
        elif report_result == "missing":
            report_result_cell.fill = yellow_fill

    # Save the workbook to a file
    wb.save(filename)
    messagebox.showinfo("Success", f"Comparison completed. Results saved to {filename}.")

def run_comparison():
    golden_json_path = golden_label.cget("text")
    report_json_path = report_label.cget("text")
    
    if not golden_json_path or not report_json_path:
        messagebox.showerror("Error", "Please upload both Golden and Report JSON files.")
    else:
        compare_jsons(golden_json_path, report_json_path)

# Tkinter UI Setup
root = tk.Tk()
root.title("JSON Comparison Tool")
root.geometry("500x400")

# Golden JSON Upload
golden_label = tk.Label(root, text="No Golden Report uploaded", wraplength=400)
golden_label.pack(pady=10)

golden_upload_btn = tk.Button(root, text="Upload Golden Report JSON", command=lambda: upload_json_file(golden_label))
golden_upload_btn.pack(pady=10)

# Report JSON Upload
report_label = tk.Label(root, text="No Report uploaded", wraplength=400)
report_label.pack(pady=10)

report_upload_btn = tk.Button(root, text="Upload Report JSON", command=lambda: upload_json_file(report_label))
report_upload_btn.pack(pady=10)

# Compare Button
compare_btn = tk.Button(root, text="Compare", command=run_comparison)
compare_btn.pack(pady=20)

root.mainloop()
